import openpyxl
import pandas as pd
from tkinter import messagebox
from sakura.utils.Lines import plot_data, read_data_from_excel

def create_summary_from_memory(forms_data, output_filepath, form5_file_path, progress_var):
    progress_var.set("Создание итогового файла")
    """
    Создает итоговый файл, собирая данные с листов "СВОД" из обработанных форм,
    и предлагает пользователю визуализировать данные с помощью графиков.

    :param forms_data: словарь, где ключи — номера форм (2, 3, 4),
                       значения — пути к обработанным файлам форм.
    :param output_filepath: путь для сохранения итогового файла.
    """
    # Создаем новый файл Excel

    summary_wb = openpyxl.Workbook()
    summary_ws = summary_wb.active
    summary_ws.title = "Итоговый СВОД"

    # Стартовая строка для копирования данных
    current_row = 1
    progress_var.set("Создание итогового файла")

    for form_number, filepath in forms_data.items():
        progress_var.set("Создание итогового файла")
        if not filepath:
            # Если файл для формы отсутствует, пропускаем её
            continue

        try:
            # Открываем обработанный файл формы

            form_wb = openpyxl.load_workbook(filepath, data_only=True)
            if "СВОД" not in form_wb.sheetnames:
                raise ValueError(f"В файле формы {form_number} отсутствует лист 'СВОД'.")

            svod_ws = form_wb["СВОД"]

            # Перед копированием данных записываем название файла
            summary_ws.cell(row=current_row, column=1, value=f"Данные из файла: {filepath}")
            current_row += 1

            # Копируем все данные с листа "СВОД"
            for row in svod_ws.iter_rows(values_only=True):
                for col_idx, cell_value in enumerate(row, start=1):
                    summary_ws.cell(row=current_row, column=col_idx, value=cell_value)
                current_row += 1

            # Добавляем пустую строку после данных текущего файла
            current_row += 1

        except Exception as e:
            print(f"Ошибка обработки файла формы {form_number}: {e}")

    # Сохраняем итоговый файл
    summary_wb.save(output_filepath)
    progress_var.set("Итоговый файл успешно создан")

    # Спрашиваем у пользователя, хочет ли он построить графики
    if messagebox.askyesno("Графики", "Хотите создать графики показателей итоговой таблицы?"):
        visualize_summary(output_filepath)
    create_ceil_model(output_filepath)
    if form5_file_path != "":
        multiply_etalons(output_filepath, form5_file_path)


def visualize_summary(file_path):

    sheet_name = "Итоговый СВОД"
    date_col = 1
    start_row = 1

    dates, values, table_titles = read_data_from_excel(file_path, sheet_name, date_col, start_row)
    plot_data(dates, values, table_titles, file_path)

def create_ceil_model(file_path):
    # Загружаем данные
    xls = pd.ExcelFile(file_path)
    sheet_name = "Итоговый СВОД"
    df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)

    # Список названий таблиц
    table_names = ["Хранение", "Входящий поток", "Исходящий поток"]

    # Список для хранения данных
    final_data = []

    # Первая таблица (стартует со 2-й строки)
    table_indices = [1]  # Индексация с нуля, поэтому 1 вместо 2

    # Поиск разделителей "Данные получены из..."
    for idx in df.index:
        if isinstance(df.iloc[idx, 0], str) and "Данные" in df.iloc[idx, 0]:
            table_indices.append(idx + 2)  # Следующая строка — начало таблицы

    # Обрабатываем найденные таблицы
    for i, start_idx in enumerate(table_indices):
        if i == len(table_names):  # Если таблиц больше, чем нужно, выходим
            break

        table_name = table_names[i]

        # Определяем конец таблицы (первая пустая строка после старта)
        end_idx = start_idx
        while end_idx < len(df) and not df.iloc[end_idx].isnull().all():
            end_idx += 1

        # Загружаем подтаблицу
        sub_df = df.iloc[start_idx-1:end_idx].copy()

        sub_df.columns = df.iloc[start_idx-1]  # Назначаем заголовки

        sub_df.dropna(how='all', axis=1, inplace=True)

        sub_df = sub_df[1:].reset_index(drop=True)  # Убираем строку с заголовками

        # Преобразуем "Объем, м3" в числа
        sub_df["Объем, м3"] = pd.to_numeric(sub_df["Объем, м3"], errors="coerce")

        # Определяем эталонный месяц (строка с максимальным "Объем, м3")
        max_month_row = sub_df.loc[sub_df["Объем, м3"].idxmax()]
        #print(max_month_row[sub_df.columns[0]])
        # Записываем данные в итоговую таблицу
        for column in sub_df.columns[1:]:

            final_data.append([table_name, max_month_row[sub_df.columns[0]], column, str(max_month_row[column]).replace('.',',')])

    # Создаем новый DataFrame
    result_df = pd.DataFrame(final_data, columns=["Тип операции","Эталонный месяц", "Показатель", "Эталонное значение"])

    # Записываем в новый лист
    with pd.ExcelWriter(file_path, mode="a", if_sheet_exists="replace") as writer:
        result_df.to_excel(writer, sheet_name="Целевая модель", index=False)

#process_excel_v2("C:\\Users\\Юрбикс\\PycharmProjects\\Sakura\\Формы\\Европрестиж\\итог.xlsx")

def multiply_etalons(file_path, form5_file_path):
    form5_path = form5_file_path.get()
    # Загружаем данные
    xls = pd.ExcelFile(file_path)
    sheet_name = "Целевая модель"
    df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
    xls = pd.ExcelFile(form5_path)
    sheet_name = "Коэффициенты роста"
    df_koef = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)

    # перевод эталонных значений в числовые (float64)
    df["Эталонное значение"] = df["Эталонное значение"].str.replace(",", ".")
    df["Эталонное значение"] = pd.to_numeric(df["Эталонное значение"], errors="coerce")

    #расчет коэффициентов роста
    for col in df_koef.columns[1:]:  # df.columns[1:] — все столбцы, кроме первого
        df_koef[col] = pd.to_numeric(df_koef[col], errors="ignore")
    #df_koef["Коэффициент роста"] = 0
    koef =[]
    for ind, row in df_koef.iterrows():
        s = 1
        for col in df_koef.columns[1:]:
            s*=row[col]
        koef.append(s)
    df_koef.insert(len(df_koef.columns), "Коэффициент роста", koef)

    #Привязываем коэффициенты роста к типам операции
    koef_dict = {"Хранение": 0, "Входящий поток": 0, "Исходящий поток": 0}
    for ind, row in df_koef.iterrows():
        if df_koef["Критерий"].iloc[ind].find("рузооборот") != -1:
            koef_dict["Входящий поток"] = df_koef["Коэффициент роста"].iloc[ind]
            koef_dict["Исходящий поток"] = df_koef["Коэффициент роста"].iloc[ind]
        elif df_koef["Критерий"].iloc[ind].find("хранения") != -1:
            koef_dict["Хранение"] = df_koef["Коэффициент роста"].iloc[ind]

    itog_koef = []

    for ind, row in df.iterrows():
        if df["Тип операции"].iloc[ind] == 'Хранение': itog_koef.append(koef_dict["Хранение"])
        if df["Тип операции"].iloc[ind] == 'Входящий поток': itog_koef.append(koef_dict["Входящий поток"])
        if df["Тип операции"].iloc[ind] == 'Исходящий поток': itog_koef.append(koef_dict["Исходящий поток"])
    df.insert(len(df.columns), "Коэффициент роста", itog_koef)

    df["Целевое значение"] = df["Эталонное значение"]*df["Коэффициент роста"]

    with pd.ExcelWriter(file_path, mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name="Целевая модель", index=False)