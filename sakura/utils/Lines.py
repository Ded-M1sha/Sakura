import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import os
import re
import numpy as np
from tkinter import messagebox


def read_data_from_excel(file_path, sheet_name, date_col, start_row):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    dates = []
    values = []
    table_titles = []
    table_start_row = None
    columns_names = []

    for row in range(start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=date_col).value

        if cell_value == "Приведенная дата":
            if table_start_row is not None:
                table_data = process_table(sheet, table_start_row + 1, row - 1, date_col, columns_names)
                dates.append(table_data['Date'].astype(str))  # Оставляем даты как строки
                values.append(table_data.drop(columns=["Date"]))

            title_row = row - 1
            title_text = sheet.cell(row=title_row, column=date_col).value
            table_titles.append(extract_table_title(title_text))
            table_start_row = row

            columns_names = [sheet.cell(row=row, column=col).value for col in range(date_col + 1, sheet.max_column + 1)
                             if sheet.cell(row=row, column=col).value]

    if table_start_row is not None:
        table_data = process_table(sheet, table_start_row + 1, sheet.max_row, date_col, columns_names)
        dates.append(table_data['Date'].astype(str))
        values.append(table_data.drop(columns=["Date"]))

    return dates, values, table_titles


def process_table(sheet, start_row, end_row, date_col, columns_names):
    dates = [str(sheet.cell(row=row, column=date_col).value) for row in
             range(start_row, end_row + 1)]  # Сохраняем даты как строки

    values = []
    for col in range(date_col + 1, sheet.max_column + 1):
        col_values = [sheet.cell(row=row, column=col).value for row in range(start_row, end_row + 1)]
        if any(val is not None for val in col_values):
            values.append(col_values)

    df = pd.DataFrame(values).T
    df.columns = columns_names
    df['Date'] = dates  # Дата теперь просто строка
    return df


def extract_table_title(title_text):
    match = re.search(r"форма (\d+)_обработанная", str(title_text), re.IGNORECASE)
    if match:
        return {"2": "Хранение", "3": "Входящий поток", "4": "Исходящий поток"}.get(match.group(1), "Неизвестная форма")
    return "Неизвестная форма"



def plot_data(dates, values, table_titles, file_path):
    output_dir = os.path.dirname(file_path)

    for date_data, value_data, table_title in zip(dates, values, table_titles):
        for column in value_data.columns:
            plt.figure(figsize=(12, 7))  # Увеличиваем размер графика

            # Создаем числовые индексы для строковых дат
            x_indices = np.arange(len(date_data))  # Числовые индексы (0, 1, 2, ...)

            # Фильтрация значений дат, чтобы оставить только те, которые заканчиваются на "20"
            filtered_dates = [date for date in date_data if date[-4: -2] == '20']
            filtered_indices = [i for i, date in enumerate(date_data) if date[-4: -2] == '20']
            filtered_values = value_data.iloc[filtered_indices]

            # Если фильтрация вернула пустой список, то продолжаем с пустым графиком
            if filtered_dates:
                plt.plot(filtered_indices, filtered_values[column], marker='o', linestyle='-', color='#e80b16', label=column)

                # Добавляем метки данных на точки графика
                for x, y in zip(filtered_indices, filtered_values[column]):
                    if isinstance(y, (int, float)) and not np.isnan(y):  # Проверяем, что значение числовое
                        plt.annotate(f'{y:.0f}', (x, y), textcoords="offset points", xytext=(0, 5),
                                     ha='center', fontsize=9, color='black',
                                     bbox=dict(boxstyle="round,pad=0.2", fc="white", ec="black", lw=0.3))

                # Находим индекс максимального значения
                if filtered_values[column].dropna().empty:  # Если столбец пуст, пропускаем
                    continue

                max_value_index = filtered_values[column].idxmax()
                max_x_index = filtered_indices[max_value_index]
                max_value = filtered_values[column].max()

                # Добавляем точку с максимальным значением
                plt.scatter(max_x_index, max_value, color='black', label=f'Эталонное значение: {max_value:.0f}', s=100, zorder=5)

                plt.title(f'{table_title}. {column}', fontsize=14, fontweight='bold', family='Calibri')
                plt.xlabel('Дата')
                plt.ylabel('Значение')
                plt.xticks(filtered_indices, filtered_dates, rotation=45)  # Устанавливаем подписи для оси X
                plt.grid(True)
                plt.legend()

                # Явно настраиваем отступы
                plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.2)

                # Сохраняем график
                output_path = os.path.join(output_dir, f'{table_title}_{column}.png')
                plt.savefig(output_path, format='png', dpi=300)
                plt.close()  # Закрываем фигуру


