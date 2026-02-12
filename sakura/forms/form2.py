import pandas as pd
import os
from openpyxl import load_workbook
from tkinter import messagebox, Toplevel, Checkbutton, IntVar, Label, Button
from datetime import datetime
from .form1 import show_error
import customtkinter as ctk

# Словарь перевода месяцев
MONTHS_RU = {
    "January": "январь", "February": "февраль", "March": "март",
    "April": "апрель", "May": "май", "June": "июнь",
    "July": "июль", "August": "август", "September": "сентябрь",
    "October": "октябрь", "November": "ноябрь", "December": "декабрь"
}

def translate_month(date):
    if pd.isnull(date):
        return None
    english_month = date.strftime("%B")
    year = date.strftime("%Y")
    return f"{MONTHS_RU.get(english_month, english_month)} {year}"


def choose_filter_columns_and_values(df, root):
    selected_filters = {}

    def select_columns():
        for col, var in column_vars.items():
            if var.get() == 1:
                selected_columns.append(col)
        column_window.destroy()

    def select_values_for_column(col):
        value_vars = {val: IntVar() for val in unique_values[col]}

        def apply_values():
            selected_filters[col] = [val for val, var in value_vars.items() if var.get() == 1]
            value_window.destroy()

        value_window = ctk.CTkToplevel(root)
        value_window.title(f"Выбор значений для {col}")
        ctk.CTkLabel(value_window, text=f"Выберите значения для фильтрации столбца '{col}':").pack()

        for val, var in value_vars.items():
            ctk.CTkCheckBox(value_window, text=val, variable=var).pack(anchor="w")

        ctk.CTkButton(value_window, text="Применить", command=apply_values).pack()
        root.wait_window(value_window)
    selected_columns = []
    column_window = ctk.CTkToplevel(root)
    column_window.title("Выбор столбцов для фильтрации")
    ctk.CTkLabel(column_window, text="Выберите столбцы для фильтрации данных:").pack()

    column_vars = {}
    columns = [col for col in df.columns if col not in ["Код товара", "Длина, см", "Ширина, см", "Высота, см"]]

    for col in columns:
        column_vars[col] = IntVar()
        ctk.CTkCheckBox(column_window, text=col, variable=column_vars[col]).pack(anchor="w")

    ctk.CTkButton(column_window, text="Далее", command=select_columns).pack()
    root.wait_window(column_window)

    unique_values = {col: df[col].dropna().unique() for col in selected_columns}

    for col in selected_columns:
        select_values_for_column(col)

    return selected_filters

def process_form2(filepath, form1_filepath, progress_var, root, on_form2_done):
    new_filepath = os.path.join(os.path.dirname(filepath), "Форма 2_обработанная.xlsx")


    progress_var.set("Загрузка данных формы 2")
    root.update()
    df_form2 = pd.read_excel(filepath)

    progress_var.set("Загрузка данных формы 1")
    root.update()
    workbook_form1 = load_workbook(form1_filepath, data_only=True)
    sheet_form1 = workbook_form1.active

    df_form1 = pd.read_excel(form1_filepath)

    replacement_value = sheet_form1["V6"].value

    if replacement_value is None:
        replacement_value = df_form1['Объем единицы итоговый, м3'].median()
        show_error("Предупреждение","Недостающие в справочнике значения будут заменены на среднее")


    if 'Код товара' not in df_form2.columns:
        show_error("Ошибка", "Отсутствует столбец 'Код товара' в форме 2. Обработка остановлена")
        root.quit()
        return


    progress_var.set("Создание словаря объемов из формы 1")
    root.update()
    volume_dict = df_form1.set_index('Код товара')['Объем единицы итоговый, м3'].to_dict()

    progress_var.set("Вычисление объемов и количеств")
    root.update()
    df_form2['Объем единицы, м3'] = df_form2['Код товара'].apply(
        lambda code: volume_dict.get(code, replacement_value)
    )
    if 'ед. изм.' in df_form2.columns:
        df_form2['Количество с учетом единицы измерения'] = df_form2.apply(
            lambda row: row['Количество'] if row['ед. изм.'] == 'шт' else (row['Количество'] // 1 + 1), axis=1
        )
    else:
        df_form2['Количество с учетом единицы измерения'] = df_form2['Количество']

    df_form2['Итоговый объем, м3'] = df_form2['Объем единицы, м3'] * df_form2['Количество с учетом единицы измерения']

    # Приводим дату к английскому формату
    df_form2['Приведенная дата'] = df_form2['Дата'].apply(
        lambda x: x.strftime("%B %Y") if pd.notnull(x) else None
    )

    df_form2.to_excel(new_filepath, index=False)

    progress_var.set("Запрос фильтров у пользователя")
    root.update()
    filter_columns_values = choose_filter_columns_and_values(df_form2, root)

    if filter_columns_values:
        for col, values in filter_columns_values.items():
            df_form2 = df_form2[df_form2[col].isin(values)]

    progress_var.set("Создание сводной таблицы")
    root.update()
    workbook = load_workbook(new_filepath)
    sheet_svod = workbook.create_sheet("СВОД")
    sheet_svod["A1"], sheet_svod["B1"], sheet_svod["C1"], sheet_svod["D1"] = \
        "Приведенная дата", "Объем, м3", "Количество уникальных номенклатур", "Количество штук"

    # Сортируем даты в хронологическом порядке
    unique_dates = sorted(
        df_form2['Приведенная дата'].dropna().unique(),
        key=lambda x: datetime.strptime(x, "%B %Y")
    )

    # Перевод названий месяцев
    def translate_month(date_str):
        english_month, year = date_str.split()
        russian_month = MONTHS_RU.get(english_month, english_month)
        return f"{russian_month} {year}"

    translated_dates = {date: translate_month(date) for date in unique_dates}

    for idx, date in enumerate(unique_dates, start=2):
        sheet_svod[f"A{idx}"] = translated_dates[date]
        date_data = df_form2[df_form2['Приведенная дата'] == date]
        sheet_svod[f"B{idx}"] = date_data['Итоговый объем, м3'].sum()
        sheet_svod[f"C{idx}"] = date_data['Код товара'].nunique()
        sheet_svod[f"D{idx}"] = date_data['Количество с учетом единицы измерения'].sum()

        progress_var.set(f"Обработка данных: {idx - 1} / {len(unique_dates)} дат")
        root.update()

    filter_row = len(unique_dates) + 3
    sheet_svod[f"A{filter_row}"] = "Параметры фильтрации:"
    for idx, (col, values) in enumerate(filter_columns_values.items(), start=filter_row + 1):
        sheet_svod[f"A{idx}"] = f"{col}: {', '.join(map(str, values))}"

    try:
        workbook.save(new_filepath)
        workbook.close()
    except Exception as e:
        messagebox.showerror("Ошибка сохранения", f"Не удалось сохранить файл. Ошибка: {e}")
        return

    show_error("Успешно", f"Файл формы 2 успешно обработан и сохранен по пути: {new_filepath}")
    on_form2_done(new_filepath)