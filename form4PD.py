import pandas as pd
import os
from openpyxl import load_workbook
from tkinter import messagebox, Toplevel, Checkbutton, IntVar, Label, Button
from datetime import datetime

# Словарь перевода месяцев
MONTHS_RU = {
    "January": "январь", "February": "февраль", "March": "март",
    "April": "апрель", "May": "май", "June": "июнь",
    "July": "июль", "August": "август", "September": "сентябрь",
    "October": "октябрь", "November": "ноябрь", "December": "декабрь"
}

def translate_month(date):
    if pd.isnull(date):
        return None
    english_month = date.strftime("%B")
    year = date.strftime("%Y")
    return f"{MONTHS_RU.get(english_month, english_month)} {year}"


# Словарь для количества дней в каждом месяце
days_in_month = {
    1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31,
    8: 31, 9: 30, 10: 31, 11: 30, 12: 31
}

def choose_filter_columns_and_values(df, root):
    """
    Функция для выбора пользователем столбцов и уникальных значений фильтрации.
    """
    selected_filters = {}

    # Первый шаг - выбор столбцов
    def select_columns():
        for col, var in column_vars.items():
            if var.get() == 1:
                selected_columns.append(col)
        column_window.destroy()

    # Второй шаг - выбор значений для каждого выбранного столбца
    def select_values_for_column(col):
        value_vars = {val: IntVar() for val in unique_values[col]}

        def apply_values():
            selected_filters[col] = [val for val, var in value_vars.items() if var.get() == 1]
            value_window.destroy()

        value_window = Toplevel(root)
        value_window.title(f"Выбор значений для {col}")
        Label(value_window, text=f"Выберите значения для фильтрации столбца '{col}':").pack()

        for val, var in value_vars.items():
            Checkbutton(value_window, text=val, variable=var).pack(anchor="w")

        Button(value_window, text="Применить", command=apply_values).pack()
        root.wait_window(value_window)

    # Окно для выбора столбцов
    selected_columns = []
    column_window = Toplevel(root)
    column_window.title("Выбор столбцов для фильтрации")
    Label(column_window, text="Выберите столбцы для фильтрации данных:").pack()

    column_vars = {}
    columns = [col for col in df.columns if col not in ["Код товара", "Длина, см", "Ширина, см", "Высота, см"]]

    for col in columns:
        column_vars[col] = IntVar()
        Checkbutton(column_window, text=col, variable=column_vars[col]).pack(anchor="w")

    Button(column_window, text="Далее", command=select_columns).pack()
    root.wait_window(column_window)

    # Получаем уникальные значения для каждого выбранного столбца
    unique_values = {col: df[col].dropna().unique() for col in selected_columns}

    # Запрашиваем выбор значений для каждого выбранного столбца
    for col in selected_columns:
        select_values_for_column(col)

    return selected_filters

def process_form4(filepath, form1_filepath, progress_var, root, on_form4_done):
    new_filepath = os.path.join(os.path.dirname(filepath), "Форма 4_обработанная.xlsx")

    try:
        # Сообщаем пользователю о загрузке файлов
        progress_var.set("Загрузка файлов...")
        root.update()

        # Загружаем данные форм
        df_form1 = pd.read_excel(form1_filepath)
        df_form4 = pd.read_excel(filepath)
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось загрузить файлы. Ошибка: {e}")
        root.quit()
        return

    # Проверка столбцов
    if 'Код товара' not in df_form4.columns or 'Код товара' not in df_form1.columns:
        messagebox.showerror("Ошибка", "Отсутствует столбец 'Код товара' в одной из форм.")
        root.quit()
        return

    # Получаем значение замены из формы 1, ячейка Q6
    workbook_form1 = load_workbook(form1_filepath, data_only=True)
    sheet_form1 = workbook_form1.active
    replacement_value = sheet_form1["Q6"].value

    if replacement_value is None:
        messagebox.showerror("Ошибка", "Отсутствует значение для замены в ячейке Q6 формы 1.")
        root.quit()
        return

    # Создаем словарь для поиска объема единицы по коду товара
    volume_dict = df_form1.set_index('Код товара')['Объем единицы после обработки выбросов, м3'].to_dict()

    # Сообщаем пользователю о добавлении новых столбцов
    progress_var.set("Добавление необходимых столбцов в форму 3...")
    root.update()

    # Добавляем столбцы в DataFrame
    df_form4['Объем единицы, м3'] = None
    df_form4['Количество с учетом единицы измерения'] = None
    df_form4['Итоговый объем, м3'] = None
    df_form4['Приведенная дата'] = None

    # Обработка строк формы 3
    progress_var.set("Обработка строк формы 3...")
    root.update()
    for idx, row in df_form4.iterrows():
        code = row['Код товара']
        date = pd.to_datetime(row['Дата'], errors='coerce', dayfirst=True)

        # Определяем "Объем единицы, м3"
        volume = volume_dict.get(code, replacement_value)

        # Приводим дату к формату "01.MM.ГГГГ" или оставляем None, если дата некорректна
        formatted_date = f"01.{date.month:02}.{date.year}" if pd.notnull(date) else None

        # Определяем "Количество с учетом единицы измерения"
        quantity = row['Количество']
        adjusted_quantity = quantity

        # Вычисляем "Итоговый объем, м3"
        final_volume = volume * adjusted_quantity

        # Записываем результаты в DataFrame
        df_form4.at[idx, 'Объем единицы, м3'] = volume
        df_form4.at[idx, 'Количество с учетом единицы измерения'] = adjusted_quantity
        df_form4.at[idx, 'Итоговый объем, м3'] = final_volume
        df_form4.at[idx, 'Приведенная дата'] = formatted_date

        # Обновление прогресса
        progress_var.set(f"Обработка данных: {idx + 1} / {len(df_form4)} строк")
        root.update()

    # Сохранение промежуточного файла
    progress_var.set("Сохранение промежуточного файла с вычисленными значениями...")
    root.update()
    df_form4.to_excel(new_filepath, index=False)

    # Запрашиваем фильтры у пользователя
    progress_var.set("Запрос фильтров у пользователя")
    root.update()
    filter_columns_values = choose_filter_columns_and_values(df_form4, root)

    # Применение фильтров
    if filter_columns_values:
        for col, values in filter_columns_values.items():
            df_form4 = df_form4[df_form4[col].isin(values)]

    # Создание листа "СВОД" и заполнение данными
    progress_var.set("Создание листа 'СВОД'...")
    root.update()

    workbook = load_workbook(new_filepath)
    sheet_svod = workbook.create_sheet("СВОД")
    sheet_svod["A1"], sheet_svod["B1"], sheet_svod["C1"], sheet_svod["D1"], sheet_svod["E1"], sheet_svod["F1"], sheet_svod["G1"], sheet_svod["H1"] = \
        "Приведенная дата", "Объем, м3", "Количество строк, штук", "Количество штук", "Количество документов, штук", "Среднесуточное количество документов", "Среднесуточный объем, м3", "Среднесуточное количество товара, штук"

    # Получаем уникальные приведенные даты
    unique_dates = sorted(df_form4['Приведенная дата'].dropna().unique())
    for idx, date in enumerate(unique_dates, start=2):
        sheet_svod[f"A{idx}"].value = date

        # Фильтруем данные по текущей дате
        filtered_data = df_form4[df_form4['Приведенная дата'] == date]

        # Заполняем данные по каждому параметру
        sheet_svod[f"B{idx}"] = filtered_data['Итоговый объем, м3'].sum()
        sheet_svod[f"C{idx}"] = filtered_data.shape[0]  # Количество строк
        sheet_svod[f"D{idx}"] = filtered_data['Количество с учетом единицы измерения'].sum()
        document_count = filtered_data['Номер документа'].nunique()
        sheet_svod[f"E{idx}"] = document_count  # Количество документов

        # Расчет среднесуточного количества документов
        month = int(date[3:5])  # Получаем месяц из даты
        daily_document_average = document_count / days_in_month[month]
        sheet_svod[f"F{idx}"] = round(daily_document_average, 2)
        daily_volume_avarage = sheet_svod[f"B{idx}"].value / days_in_month[month]
        sheet_svod[f"G{idx}"] = round(daily_volume_avarage, 2)
        daily_quantity_avarage = sheet_svod[f"D{idx}"].value / days_in_month[month]
        sheet_svod[f"H{idx}"] = round(daily_quantity_avarage, 2)
        # Обновление прогресса
        progress_var.set(f"Заполнение сводных данных: {idx - 1} / {len(unique_dates)} дат")
        root.update()

    # Запись параметров фильтрации под таблицей
    filter_row = len(unique_dates) + 3
    sheet_svod[f"A{filter_row}"] = "Параметры фильтрации:"
    for idx, (col, values) in enumerate(filter_columns_values.items(), start=filter_row + 1):
        sheet_svod[f"A{idx}"] = f"{col}: {', '.join(map(str, values))}"

    # Сохранение и закрытие файла
    progress_var.set("Сохранение окончательного файла формы 4...")
    root.update()
    try:
        workbook.save(new_filepath)
        workbook.close()
    except Exception as e:
        messagebox.showerror("Ошибка сохранения", f"Не удалось сохранить файл. Ошибка: {e}")
        return

    # Сообщение об успешной обработке
    messagebox.showinfo("Успешно", f"Файл формы 4 успешно обработан и сохранен по пути: {new_filepath}")
    on_form4_done(new_filepath)