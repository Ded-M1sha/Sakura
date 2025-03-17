import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import re  # Для извлечения числа из строки

import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import os
import re
from tkinter import messagebox


def read_data_from_excel(file_path, sheet_name, date_col, start_row):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    dates = []
    values = []
    table_titles = []
    table_start_row = None
    columns_names = []

    for row in range(start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=date_col).value

        if cell_value == "Приведенная дата":
            if table_start_row is not None:
                table_data = process_table(sheet, table_start_row + 1, row - 1, date_col, columns_names)
                dates.append(table_data['Date'].astype(str))  # Оставляем даты как строки
                values.append(table_data.drop(columns=["Date"]))

            title_row = row - 1
            title_text = sheet.cell(row=title_row, column=date_col).value
            table_titles.append(extract_table_title(title_text))
            table_start_row = row

            columns_names = [sheet.cell(row=row, column=col).value for col in range(date_col + 1, sheet.max_column + 1)
                             if sheet.cell(row=row, column=col).value]

    if table_start_row is not None:
        table_data = process_table(sheet, table_start_row + 1, sheet.max_row, date_col, columns_names)
        dates.append(table_data['Date'].astype(str))
        values.append(table_data.drop(columns=["Date"]))

    return dates, values, table_titles


def process_table(sheet, start_row, end_row, date_col, columns_names):
    dates = [str(sheet.cell(row=row, column=date_col).value) for row in
             range(start_row, end_row + 1)]  # Сохраняем даты как строки

    values = []
    for col in range(date_col + 1, sheet.max_column + 1):
        col_values = [sheet.cell(row=row, column=col).value for row in range(start_row, end_row + 1)]
        if any(val is not None for val in col_values):
            values.append(col_values)

    df = pd.DataFrame(values).T
    df.columns = columns_names
    df['Date'] = dates  # Дата теперь просто строка
    return df


def extract_table_title(title_text):
    match = re.search(r"форма (\d+)_обработанная", str(title_text), re.IGNORECASE)
    if match:
        return {"2": "Хранение", "3": "Входящий поток", "4": "Исходящий поток"}.get(match.group(1), "Неизвестная форма")
    return "Неизвестная форма"


def plot_data(dates, values, table_titles, file_path):
    output_dir = os.path.dirname(file_path)

    for date_data, value_data, table_title in zip(dates, values, table_titles):
        for column in value_data.columns:
            plt.figure(figsize=(10, 6))
            plt.plot(date_data, value_data[column], marker='o', linestyle='-', color='#e80b16', label=column)

            plt.title(f'{table_title}. {column}', fontsize=14, fontweight='bold', family='Arial')
            plt.xlabel('Дата')
            plt.ylabel('Значение')
            plt.xticks(rotation=45)
            plt.grid(True)
            plt.legend()
            plt.tight_layout()

            output_path = os.path.join(output_dir, f'{table_title}_{column}.png')
            plt.savefig(output_path, format='png', dpi=300)
            plt.show()


