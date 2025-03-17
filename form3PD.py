import pandas as pd
import os
from openpyxl import load_workbook
from tkinter import messagebox, Toplevel, Checkbutton, IntVar, Label, Button
from openpyxl.utils import get_column_letter
from datetime import datetime

# Словарь перевода месяцев
MONTHS_RU = {
    "January": "январь", "February": "февраль", "March": "март",
    "April": "апрель", "May": "май", "June": "июнь",
    "July": "июль", "August": "август", "September": "сентябрь",
    "October": "октябрь", "November": "ноябрь", "December": "декабрь"
}

def translate_month(date):
    if pd.isnull(date):
        return None
    english_month = date.strftime("%B")
    year = date.strftime("%Y")
    return f"{MONTHS_RU.get(english_month, english_month)} {year}"


# Словарь для количества дней в каждом месяце
days_in_month = {
    1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31,
    8: 31, 9: 30, 10: 31, 11: 30, 12: 31
}
def choose_filter_columns_and_values(df, root):
    """
    Функция для выбора пользователем столбцов и уникальных значений фильтрации.
    """
    selected_filters = {}

    # Первый шаг - выбор столбцов
    def select_columns():
        for col, var in column_vars.items():
            if var.get() == 1:
                selected_columns.append(col)
        column_window.destroy()

    # Второй шаг - выбор значений для каждого выбранного столбца
    def select_values_for_column(col):
        value_vars = {val: IntVar() for val in unique_values[col]}

        def apply_values():
            selected_filters[col] = [val for val, var in value_vars.items() if var.get() == 1]
            value_window.destroy()

        value_window = Toplevel(root)
        value_window.title(f"Выбор значений для {col}")
        Label(value_window, text=f"Выберите значения для фильтрации столбца '{col}':").pack()

        for val, var in value_vars.items():
            Checkbutton(value_window, text=val, variable=var).pack(anchor="w")

        Button(value_window, text="Применить", command=apply_values).pack()
        root.wait_window(value_window)

    # Окно для выбора столбцов
    selected_columns = []
    column_window = Toplevel(root)
    column_window.title("Выбор столбцов для фильтрации")
    Label(column_window, text="Выберите столбцы для фильтрации данных:").pack()

    column_vars = {}
    columns = [col for col in df.columns if col not in ["Код товара", "Длина, см", "Ширина, см", "Высота, см"]]

    for col in columns:
        column_vars[col] = IntVar()
        Checkbutton(column_window, text=col, variable=column_vars[col]).pack(anchor="w")

    Button(column_window, text="Далее", command=select_columns).pack()
    root.wait_window(column_window)

    # Получаем уникальные значения для каждого выбранного столбца
    unique_values = {col: df[col].dropna().unique() for col in selected_columns}

    # Запрашиваем выбор значений для каждого выбранного столбца
    for col in selected_columns:
        select_values_for_column(col)

    return selected_filters

def process_form3(filepath, form1_filepath, progress_var, root, on_form3_done):
    new_filepath = os.path.join(os.path.dirname(filepath), "Форма 3_обработанная.xlsx")

    try:
        # Сообщаем пользователю о загрузке файлов
        progress_var.set("Загрузка файлов...")
        root.update()

        # Загружаем данные форм
        df_form1 = pd.read_excel(form1_filepath)
        df_form3 = pd.read_excel(filepath)
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось загрузить файлы. Ошибка: {e}")
        root.quit()
        return

    # Проверка столбцов
    if 'Код товара' not in df_form3.columns or 'Код товара' not in df_form1.columns:
        messagebox.showerror("Ошибка", "Отсутствует столбец 'Код товара' в одной из форм.")
        root.quit()
        return

    # Получаем значение замены из формы 1, ячейка Q6
    workbook_form1 = load_workbook(form1_filepath, data_only=True)
    sheet_form1 = workbook_form1.active

    # Сохраняем значение ячейки Q6 как переменную
    replacement_value = sheet_form1["V6"].value

    # Проверяем, что значение не пустое
    if replacement_value is None:
        messagebox.showerror("Ошибка", "Отсутствует значение для замены в ячейке V6 формы 1.")
        root.quit()
        return

    # Создаем словарь для поиска объема единицы по коду товара
    volume_dict = df_form1.set_index('Код товара')['Объем единицы итоговый, м3'].to_dict()

    # Сообщаем пользователю о добавлении новых столбцов
    progress_var.set("Добавление необходимых столбцов в форму 3...")
    root.update()

    # Добавляем столбцы в DataFrame
    df_form3['Объем единицы, м3'] = None
    df_form3['Количество с учетом единицы измерения'] = None
    df_form3['Итоговый объем, м3'] = None
    df_form3['Приведенная дата'] = None

    # Обработка строк формы 3
    progress_var.set("Обработка строк формы 3...")
    root.update()
    # Приводим дату к английскому формату

    for idx, row in df_form3.iterrows():
        code = row['Код товара']
        date = pd.to_datetime(row['Дата'], errors='coerce', dayfirst=True)

        # Определяем "Объем единицы, м3"
        volume = volume_dict.get(code, replacement_value)

        # Приводим дату к формату "01.MM.ГГГГ" или оставляем None, если дата некорректна
        #formatted_date = f"01.{date.month:02}.{date.year}" if pd.notnull(date) else None



        # Определяем "Количество с учетом единицы измерения"
        # quantity = row['Количество']
        # if row['ед. изм.'] in df_form3:
        #     unit = row['ед. изм.']
        #     adjusted_quantity = quantity if unit == 'шт' else (quantity // 1)
        #
        # else:
        #     adjusted_quantity = quantity



        # Вычисляем "Итоговый объем, м3"
        final_volume = volume * row['Количество']

        # Записываем результаты в DataFrame
        df_form3.at[idx, 'Объем единицы, м3'] = volume

        df_form3.at[idx, 'Итоговый объем, м3'] = final_volume


        # Обновление прогресса
        progress_var.set(f"Обработка данных: {idx + 1} / {len(df_form3)} строк")
        root.update()
    print(type(date))

    # Приведение столбца 'Дата' к формату datetime
    df_form3['Дата'] = pd.to_datetime(df_form3['Дата'], errors='coerce', dayfirst=True)

    # Создание колонки 'Приведенная дата' с отформатированной датой
    df_form3['Приведенная дата'] = df_form3['Дата'].apply(
        lambda x: x.strftime("%B %Y") if pd.notnull(x) else None
    )

    # Сохранение промежуточного файла
    progress_var.set("Сохранение промежуточного файла с вычисленными значениями...")
    root.update()
    df_form3.to_excel(new_filepath, index=False)

    # Запрашиваем фильтры у пользователя
    progress_var.set("Запрос фильтров у пользователя")
    root.update()
    filter_columns_values = choose_filter_columns_and_values(df_form3, root)

    # Применение фильтров
    if filter_columns_values:
        for col, values in filter_columns_values.items():
            df_form3 = df_form3[df_form3[col].isin(values)]



    # Создание листа "СВОД" и заполнение данными
    progress_var.set("Создание листа 'СВОД'...")
    root.update()

    workbook = load_workbook(new_filepath)
    sheet_svod = workbook.create_sheet("СВОД")
    sheet_svod["A1"], sheet_svod["B1"], sheet_svod["C1"], sheet_svod["D1"], sheet_svod["E1"], sheet_svod["F1"], sheet_svod["G1"], sheet_svod["H1"] = \
        "Приведенная дата", "Объем, м3", "Количество строк", "Количество штук", "Количество документов, штук", "Среднесуточное количество документов", "Среднесуточный объем, м3", "Среднесуточное количество товара, штук"

    # Сортируем даты в хронологическом порядке
    unique_dates = sorted(
        df_form3['Приведенная дата'].dropna().unique(),
        key=lambda x: datetime.strptime(x, "%B %Y")
    )

    # Перевод названий месяцев
    def translate_month(date_str):
        try:
            english_month, year = date_str.split()
            russian_month = MONTHS_RU.get(english_month, english_month)
            return f"{russian_month} {year}"
        except Exception as e:
            print(f"Ошибка перевода месяца для {date_str}: {e}")
            return date_str

    translated_dates = {date: translate_month(date) for date in unique_dates}

    # Обработка данных для листа "СВОД"
    for idx, date in enumerate(unique_dates, start=2):
        sheet_svod[f"A{idx}"] = translated_dates[date]
        date_data = df_form3[df_form3['Приведенная дата'] == date]

        try:
            # Предполагаем, что формат даты: "January 2023"
            month_str = date.split()[0]
            month = datetime.strptime(month_str, "%B").month
        except (ValueError, AttributeError) as e:
            print(f"Ошибка обработки даты: {date}, ошибка: {e}")
            continue  # Пропускаем некорректные даты

        # Заполняем данные
        sheet_svod[f"B{idx}"] = date_data['Итоговый объем, м3'].sum()
        sheet_svod[f"C{idx}"] = date_data.shape[0]  # Количество строк
        sheet_svod[f"D{idx}"] = date_data['Количество'].sum()
        document_count = date_data['Номер документа'].nunique()
        sheet_svod[f"E{idx}"] = document_count

        if month is not None:
            daily_document_average = document_count / days_in_month[month]
            daily_volume_average = sheet_svod[f"B{idx}"].value / days_in_month[month]
            daily_quantity_average = sheet_svod[f"D{idx}"].value / days_in_month[month]

            sheet_svod[f"F{idx}"] = round(daily_document_average, 2)
            sheet_svod[f"G{idx}"] = round(daily_volume_average, 2)
            sheet_svod[f"H{idx}"] = round(daily_quantity_average, 2)

        # Обновление прогресса
        progress_var.set(f"Заполнение сводных данных: {idx - 1} / {len(unique_dates)} дат")
        root.update()

    # Запись параметров фильтрации под таблицей
    filter_row = len(unique_dates) + 3
    sheet_svod[f"A{filter_row}"] = "Параметры фильтрации:"
    for idx, (col, values) in enumerate(filter_columns_values.items(), start=filter_row + 1):
        sheet_svod[f"A{idx}"] = f"{col}: {', '.join(map(str, values))}"

    # Сохранение и закрытие файла
    progress_var.set("Сохранение окончательного файла формы 3...")
    root.update()
    try:
        workbook.save(new_filepath)
        workbook.close()
    except Exception as e:
        messagebox.showerror("Ошибка сохранения", f"Не удалось сохранить файл. Ошибка: {e}")
        print(e)
        return

    # Сообщение об успешной обработке
    messagebox.showinfo("Успешно", f"Файл формы 3 успешно обработан и сохранен по пути: {new_filepath}")
    on_form3_done(new_filepath)