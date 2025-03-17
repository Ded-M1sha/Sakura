import openpyxl

from tkinter import messagebox
from Lines import plot_data, read_data_from_excel

def create_summary_from_memory(forms_data, output_filepath):
    """
    Создает итоговый файл, собирая данные с листов "СВОД" из обработанных форм,
    и предлагает пользователю визуализировать данные с помощью графиков.

    :param forms_data: словарь, где ключи — номера форм (2, 3, 4),
                       значения — пути к обработанным файлам форм.
    :param output_filepath: путь для сохранения итогового файла.
    """
    # Создаем новый файл Excel
    summary_wb = openpyxl.Workbook()
    summary_ws = summary_wb.active
    summary_ws.title = "Итоговый СВОД"

    # Стартовая строка для копирования данных
    current_row = 1

    for form_number, filepath in forms_data.items():
        if not filepath:
            # Если файл для формы отсутствует, пропускаем её
            continue

        try:
            # Открываем обработанный файл формы
            form_wb = openpyxl.load_workbook(filepath, data_only=True)
            if "СВОД" not in form_wb.sheetnames:
                raise ValueError(f"В файле формы {form_number} отсутствует лист 'СВОД'.")

            svod_ws = form_wb["СВОД"]

            # Перед копированием данных записываем название файла
            summary_ws.cell(row=current_row, column=1, value=f"Данные из файла: {filepath}")
            current_row += 1

            # Копируем все данные с листа "СВОД"
            for row in svod_ws.iter_rows(values_only=True):
                for col_idx, cell_value in enumerate(row, start=1):
                    summary_ws.cell(row=current_row, column=col_idx, value=cell_value)
                current_row += 1

            # Добавляем пустую строку после данных текущего файла
            current_row += 1

        except Exception as e:
            print(f"Ошибка обработки файла формы {form_number}: {e}")

    # Сохраняем итоговый файл
    summary_wb.save(output_filepath)
    messagebox.showinfo("Успешно", "Итоговый файл успешно создан!")

    # Спрашиваем у пользователя, хочет ли он построить графики
    if messagebox.askyesno("Графики", "Хотите ли просмотреть графики показателей итоговой таблицы?"):
        visualize_summary(output_filepath)


def visualize_summary(file_path):
    """
    Визуализирует графики из итогового файла.
    """
    sheet_name = "Итоговый СВОД"
    date_col = 1  # Номер столбца с датами (начиная с 1)
    start_row = 1

    # Чтение данных
    dates, values, table_titles = read_data_from_excel(file_path, sheet_name, date_col, start_row)

    # Построение графиков
    plot_data(dates, values, table_titles, file_path)
