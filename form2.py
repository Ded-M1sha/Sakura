import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from tkinter import messagebox
from datetime import datetime


def process_form2(filepath, form1_filepath, progress_var, root, on_form2_done):
    new_filepath = os.path.join(os.path.dirname(filepath), "Форма 2_обработанная.xlsx")

    try:
        # Загружаем оба файла
        df_form1 = pd.read_excel(form1_filepath)
        df_form2 = pd.read_excel(filepath)
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось загрузить файлы. Ошибка: {e}")
        root.quit()
        return

    # Проверяем наличие необходимых столбцов
    if 'Код товара' not in df_form1.columns or 'Код товара' not in df_form2.columns:
        messagebox.showerror("Ошибка", "Отсутствует столбец 'Код товара' в одной из форм.")
        root.quit()
        return

    # Добавляем необходимые столбцы
    df_form2['Объем единицы, м3'] = None
    df_form2['Количество с учетом единицы измерения'] = None
    df_form2['Итоговый объем, м3'] = None

    # Создаем столбец "Приведенная дата" с датами, приведенными к формату "01.ММ.ГГГГ"
    df_form2['Приведенная дата'] = df_form2['Дата'].apply(
        lambda x: f"01.{x.month:02}.{x.year}" if pd.notnull(x) else None
    )

    # Сохраняем промежуточный файл для работы с формулами
    df_form2.to_excel(new_filepath, index=False)

    # Открываем файл для работы с формулами через openpyxl
    workbook = load_workbook(new_filepath)
    sheet = workbook.active

    # Определяем индексы нужных столбцов
    code_col = get_column_letter(df_form2.columns.get_loc("Код товара") + 1)
    start_quantity_col = get_column_letter(df_form2.columns.get_loc("Количество") + 1)
    quantity_col = get_column_letter(df_form2.columns.get_loc("Количество с учетом единицы измерения") + 1)
    unit_col = get_column_letter(df_form2.columns.get_loc("ед. изм.") + 1)
    volume_col = get_column_letter(df_form2.columns.get_loc("Объем единицы, м3") + 1)
    final_volume_col = get_column_letter(df_form2.columns.get_loc("Итоговый объем, м3") + 1)
    date_col = get_column_letter(df_form2.columns.get_loc("Приведенная дата") + 1)

    # Заполняем формулы
    for row in range(2, len(df_form2) + 2):
        # Формула для столбца "Объем единицы, м3" с внешней ссылкой
        volume_formula = f'=IFERROR(VLOOKUP({code_col}{row}, \'[{os.path.basename(form1_filepath)}]Sheet1\'!A:L, 12, 0), \'[{os.path.basename(form1_filepath)}]Sheet1\'!$Q$6)'
        sheet[f"{volume_col}{row}"].value = volume_formula

        # Формула для столбца "Количество с учетом единицы измерения"
        quantity_formula = f'=IF({unit_col}{row}="шт", {start_quantity_col}{row}, CEILING({start_quantity_col}{row}, 1))'
        sheet[f"{quantity_col}{row}"].value = quantity_formula

        # Формула для столбца "Итоговый объем, м3"
        final_volume_formula = f"={volume_col}{row}*{quantity_col}{row}"
        sheet[f"{final_volume_col}{row}"].value = final_volume_formula

        # Обновление прогресса
        progress_var.set(f"Обработка данных: {row - 1} / {len(df_form2)} строк")
        root.update()

    # Создаем лист "СВОД"
    sheet_svod = workbook.create_sheet("СВОД")

    # Добавляем заголовки
    sheet_svod["A1"] = "Приведенная дата"
    sheet_svod["B1"] = "Объем, м3"
    sheet_svod["C1"] = "Количество, строк"
    sheet_svod["D1"] = "Количество, штук"

    # Получаем уникальные приведенные даты
    unique_dates = df_form2["Приведенная дата"].dropna().unique()

    # Подсчет уникальных кодов товаров и заполнение данных
    for idx, date in enumerate(sorted(unique_dates), start=2):
        sheet_svod[f"A{idx}"].value = date

        # Формула для суммы объемов по приведенной дате
        date_cell = f"A{idx}"
        sheet_svod[
            f"B{idx}"] = f"=SUMIF('{sheet.title}'!{date_col}:{date_col}, {date_cell}, '{sheet.title}'!{final_volume_col}:{final_volume_col})"

        # Подсчет количества уникальных кодов товаров на уровне Python
        unique_codes_count = df_form2[df_form2["Приведенная дата"] == date]["Код товара"].nunique()
        sheet_svod[f"C{idx}"].value = unique_codes_count

        # Формула для суммы количества с учетом единицы измерения
        sheet_svod[
            f"D{idx}"] = f"=SUMIF('{sheet.title}'!{date_col}:{date_col}, {date_cell}, '{sheet.title}'!{quantity_col}:{quantity_col})"

    # Сохранение и закрытие файла
    try:
        workbook.save(new_filepath)
        workbook.close()
    except Exception as e:
        messagebox.showerror("Ошибка сохранения", f"Не удалось сохранить файл. Ошибка: {e}")
        return

    # Сообщаем об успешной обработке
    messagebox.showinfo("Успешно", f"Файл формы 2 успешно обработан и сохранен по пути: {new_filepath}")
    #on_form2_done()
