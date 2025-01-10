import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from tkinter import messagebox

def process_form1(filepath, progress_var, root, on_form1_done, upper_limit):
    # Путь для сохранения обработанного файла
    new_filepath = os.path.join(os.path.dirname(filepath), "Форма 1_ обработанная.xlsx")

    # Загружаем данные из исходного файла
    df = pd.read_excel(filepath)

    # Проверяем наличие необходимых столбцов
    required_columns = ["Код товара", "Название товара", "Группа товаров", "Длина, см", "Ширина, см", "Высота, см"]
    for col in required_columns:
        if col not in df.columns:
            messagebox.showerror("Ошибка", f"Отсутствует необходимый столбец: {col}")
            root.quit()
            return

    # Добавляем столбцы для расчета объемов и обработки выбросов
    df['Объем единицы, м3'] = None
    df['Объем единицы после аппроксимации, м3'] = None
    df['Является ли выбросом?'] = None
    df['Объем единицы после обработки выбросов, м3'] = None

    # Сохраняем промежуточный файл для работы с формулами через openpyxl
    df.to_excel(new_filepath, index=False)

    # Работаем с формулами через openpyxl
    workbook = load_workbook(new_filepath)
    sheet = workbook.active

    # Индексы для нужных столбцов
    length_col = get_column_letter(df.columns.get_loc("Длина, см") + 1)
    width_col = get_column_letter(df.columns.get_loc("Ширина, см") + 1)
    height_col = get_column_letter(df.columns.get_loc("Высота, см") + 1)
    volume_col = get_column_letter(df.columns.get_loc("Объем единицы, м3") + 1)
    approximated_volume_col = get_column_letter(df.columns.get_loc("Объем единицы после аппроксимации, м3") + 1)
    group_col = get_column_letter(df.columns.get_loc("Группа товаров") + 1)
    is_outlier_col = get_column_letter(df.columns.get_loc("Является ли выбросом?") + 1)
    processed_volume_col = get_column_letter(df.columns.get_loc("Объем единицы после обработки выбросов, м3") + 1)

    # Добавляем текстовые метки в ячейки P1-P6
    sheet["P1"], sheet["P2"], sheet["P3"], sheet["P4"], sheet["P5"], sheet["P6"] = \
        "Q1", "Q3", "IQR", "Среднее арифметическое после аппроксимации", "Медиана после аппроксимации", "Значение для замены"

    # Формулы для ячеек Q1, Q2, Q3, Q4, Q5
    sheet["Q1"] = "=QUARTILE(J:J, 1)"  # Q1
    sheet["Q2"] = "=QUARTILE(J:J, 3)"  # Q3
    sheet["Q3"] = "=Q2 - Q1"           # IQR
    sheet["Q4"] = "=AVERAGE(J:J)"       # Среднее арифметическое
    sheet["Q5"] = "=MEDIAN(J:J)"        # Медиана

    # Условие для Q6, в зависимости от значения upper_limit
    if upper_limit == 0:
        # Если пользователь ввел 0, используем формулу сравнения медианы и среднего арифметического
        sheet["Q6"] = "=IF(ABS(Q5 - Q4) / Q4 > 0.1, Q5, Q4)"
    else:
        # Иначе используем значение upper_limit
        sheet["Q6"] = upper_limit

    # Заполняем формулы для расчета объема, аппроксимации и работы с выбросами
    for row in range(2, len(df) + 2):
        # Формула для расчета объема
        volume_formula = f"=IFERROR(IF(OR({length_col}{row}=0, {width_col}{row}=0, {height_col}{row}=0), " \
                         f"\"нет данных по объему\", {length_col}{row}*{width_col}{row}*{height_col}{row}*0.000001), " \
                         f"\"нет данных по объему\")"
        sheet[f"{volume_col}{row}"] = volume_formula

        # Формула для аппроксимации на основе товарной группы
        approximation_formula = f"=IF({volume_col}{row}=\"нет данных по объему\", " \
                                f"IFERROR(AVERAGEIF({group_col}:{group_col}, {group_col}{row}, " \
                                f"{volume_col}:{volume_col}), " \
                                f"AVERAGE({volume_col}:{volume_col})), {volume_col}{row})"
        sheet[f"{approximated_volume_col}{row}"] = approximation_formula

        # Формула для определения выбросов
        outlier_formula = f"=IF(OR({approximated_volume_col}{row}>=($Q$2 + 1.5*$Q$3), " \
                          f"{approximated_volume_col}{row}<=IF(($Q$1 - 1.5*$Q$3) <= 0, 0, ($Q$1 - 1.5*$Q$3))), " \
                          f"\"Да\", \"Нет\")"
        sheet[f"{is_outlier_col}{row}"] = outlier_formula

        # Формула для замены выбросов на значение из Q6
        processed_volume_formula = f"=IF({is_outlier_col}{row}=\"Нет\", {approximated_volume_col}{row}, $Q$6)"
        sheet[f"{processed_volume_col}{row}"] = processed_volume_formula

        # Обновляем прогресс обработки
        progress_var.set(f"Обработка данных: {row - 1} / {len(df)} строк")
        root.update()

    # Сохраняем файл с формулами
    workbook.save(new_filepath)
    workbook.close()

    # Информируем пользователя об успешном завершении и завершаем приложение
    messagebox.showinfo("Успешно", f"Файл успешно сохранен по пути: {new_filepath}")
    root.quit()

    # Вызов обработчика для продолжения работы
    #on_form1_done(new_filepath)
