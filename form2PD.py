import pandas as pd
import os
from openpyxl import load_workbook
from tkinter import messagebox, Toplevel, Checkbutton, IntVar, Label, Button


def choose_filter_columns_and_values(df, root):
    """
    Функция для выбора пользователем столбцов и уникальных значений фильтрации.
    """
    selected_filters = {}

    # Первый шаг - выбор столбцов
    def select_columns():
        for col, var in column_vars.items():
            if var.get() == 1:
                selected_columns.append(col)
        column_window.destroy()

    # Второй шаг - выбор значений для каждого выбранного столбца
    def select_values_for_column(col):
        value_vars = {val: IntVar() for val in unique_values[col]}

        def apply_values():
            selected_filters[col] = [val for val, var in value_vars.items() if var.get() == 1]
            value_window.destroy()

        value_window = Toplevel(root)
        value_window.title(f"Выбор значений для {col}")
        Label(value_window, text=f"Выберите значения для фильтрации столбца '{col}':").pack()

        for val, var in value_vars.items():
            Checkbutton(value_window, text=val, variable=var).pack(anchor="w")

        Button(value_window, text="Применить", command=apply_values).pack()
        root.wait_window(value_window)

    # Окно для выбора столбцов
    selected_columns = []
    column_window = Toplevel(root)
    column_window.title("Выбор столбцов для фильтрации")
    Label(column_window, text="Выберите столбцы для фильтрации данных:").pack()

    column_vars = {}
    columns = [col for col in df.columns if col not in ["Код товара", "Длина, см", "Ширина, см", "Высота, см"]]

    for col in columns:
        column_vars[col] = IntVar()
        Checkbutton(column_window, text=col, variable=column_vars[col]).pack(anchor="w")

    Button(column_window, text="Далее", command=select_columns).pack()
    root.wait_window(column_window)

    # Получаем уникальные значения для каждого выбранного столбца
    unique_values = {col: df[col].dropna().unique() for col in selected_columns}

    # Запрашиваем выбор значений для каждого выбранного столбца
    for col in selected_columns:
        select_values_for_column(col)

    return selected_filters


def process_form2(filepath, form1_filepath, progress_var, root, on_form2_done):
    # Определение пути для сохранения обработанного файла
    new_filepath = os.path.join(os.path.dirname(filepath), "Форма 2_обработанная.xlsx")

    try:
        # Загружаем данные формы 2
        progress_var.set("Загрузка данных формы 2")
        root.update()
        df_form2 = pd.read_excel(filepath)

        # Загружаем файл формы 1 с помощью openpyxl для доступа к конкретной ячейке
        progress_var.set("Загрузка данных формы 1")
        root.update()
        workbook_form1 = load_workbook(form1_filepath, data_only=True)
        sheet_form1 = workbook_form1.active
        replacement_value = sheet_form1["Q6"].value

        if replacement_value is None:
            raise ValueError("Ячейка Q6 в форме 1 не содержит значение для замены.")

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось загрузить файлы или прочитать значение для замены. Ошибка: {e}")
        root.quit()
        return

    # Проверка наличия необходимых столбцов
    if 'Код товара' not in df_form2.columns:
        messagebox.showerror("Ошибка", "Отсутствует столбец 'Код товара' в форме 2.")
        root.quit()
        return

    # Загружаем форму 1 для поиска объемов по коду товара
    df_form1 = pd.read_excel(form1_filepath)

    # Создаем словарь для поиска объема единицы по коду товара из формы 1
    progress_var.set("Создание словаря объемов из формы 1")
    root.update()
    volume_dict = df_form1.set_index('Код товара')['Объем единицы после аппроксимации, м3'].to_dict()

    # Добавляем столбцы для расчетов
    progress_var.set("Вычисление объемов и количеств")
    root.update()
    df_form2['Объем единицы, м3'] = df_form2['Код товара'].apply(
        lambda code: volume_dict.get(code, replacement_value)
    )
    if 'ед. изм.' in df_form2.columns:
        df_form2['Количество с учетом единицы измерения'] = df_form2.apply(
            lambda row: row['Количество'] if row['ед. изм.'] == 'шт' else (row['Количество'] // 1 + 1), axis=1
        )
    else:
        df_form2['Количество с учетом единицы измерения'] = df_form2['Количество']


    df_form2['Итоговый объем, м3'] = df_form2['Объем единицы, м3'] * df_form2['Количество с учетом единицы измерения']

    # Приводим дату к нужному формату
    df_form2['Приведенная дата'] = df_form2['Дата'].apply(
        lambda x: f"01.{x.month:02}.{x.year}" if pd.notnull(x) else None
    )

    # Сохраняем данные с новыми расчетными столбцами
    df_form2.to_excel(new_filepath, index=False)

    # Запрашиваем фильтры у пользователя
    progress_var.set("Запрос фильтров у пользователя")
    root.update()
    filter_columns_values = choose_filter_columns_and_values(df_form2, root)

    # Применение фильтров
    if filter_columns_values:
        for col, values in filter_columns_values.items():
            df_form2 = df_form2[df_form2[col].isin(values)]

    # Создаем лист "СВОД" и заполняем его данными
    progress_var.set("Создание сводной таблицы")
    root.update()
    workbook = load_workbook(new_filepath)
    sheet_svod = workbook.create_sheet("СВОД")
    sheet_svod["A1"], sheet_svod["B1"], sheet_svod["C1"], sheet_svod["D1"] = \
        "Приведенная дата", "Объем, м3", "Количество строк", "Количество штук"

    # Получаем уникальные приведенные даты и агрегацию данных по каждой дате
    unique_dates = sorted(df_form2['Приведенная дата'].dropna().unique())
    for idx, date in enumerate(unique_dates, start=2):
        sheet_svod[f"A{idx}"] = date
        date_data = df_form2[df_form2['Приведенная дата'] == date]
        sheet_svod[f"B{idx}"] = date_data['Итоговый объем, м3'].sum()
        sheet_svod[f"C{idx}"] = date_data['Код товара'].count()
        sheet_svod[f"D{idx}"] = date_data['Количество с учетом единицы измерения'].sum()

        # Обновление прогресса
        progress_var.set(f"Обработка данных: {idx - 1} / {len(unique_dates)} дат")
        root.update()

    # Запись параметров фильтрации под таблицей
    filter_row = len(unique_dates) + 3
    sheet_svod[f"A{filter_row}"] = "Параметры фильтрации:"
    for idx, (col, values) in enumerate(filter_columns_values.items(), start=filter_row + 1):
        sheet_svod[f"A{idx}"] = f"{col}: {', '.join(map(str, values))}"

    # Сохраняем и закрываем файл
    try:
        workbook.save(new_filepath)
        workbook.close()
    except Exception as e:
        messagebox.showerror("Ошибка сохранения", f"Не удалось сохранить файл. Ошибка: {e}")
        return

    # Сообщение об успешной обработке
    messagebox.showinfo("Успешно", f"Файл формы 2 успешно обработан и сохранен по пути: {new_filepath}")
    on_form2_done(new_filepath)