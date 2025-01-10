import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from tkinter import messagebox


def process_form4(filepath, form1_filepath, progress_var, root, on_form4_done):
    new_filepath = os.path.join(os.path.dirname(filepath), "Форма 4_обработанная.xlsx")

    try:
        # Загружаем оба файла
        df_form1 = pd.read_excel(form1_filepath)
        df_form4 = pd.read_excel(filepath)
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось загрузить файлы. Ошибка: {e}")
        root.quit()
        return

    # Проверка на наличие необходимых столбцов
    if 'Код товара' not in df_form4.columns or 'Код товара' not in df_form1.columns:
        messagebox.showerror("Ошибка", "Отсутствует столбец 'Код товара' в одной из форм.")
        root.quit()
        return

    # Добавляем необходимые столбцы для формы 3
    df_form4['Объем единицы, м3'] = None
    df_form4['Количество с учетом единицы измерения'] = None
    df_form4['Итоговый объем, м3'] = None
    df_form4['Приведенная дата'] = df_form4['Дата'].apply(
        lambda
            x: f"01.{pd.to_datetime(x, dayfirst=True).month:02}.{pd.to_datetime(x, dayfirst=True).year}" if pd.notnull(
            x) else None
    )

    # Сохраняем промежуточный файл для работы с формулами через openpyxl
    df_form4.to_excel(new_filepath, index=False)

    # Открываем файл для работы с формулами через openpyxl
    workbook = load_workbook(new_filepath)
    sheet = workbook.active

    # Определяем индексы нужных столбцов
    code_col = get_column_letter(df_form4.columns.get_loc("Код товара") + 1)
    start_quantity_col = get_column_letter(df_form4.columns.get_loc("Количество") + 1)
    quantity_col = get_column_letter(df_form4.columns.get_loc("Количество с учетом единицы измерения") + 1)
    unit_col = get_column_letter(df_form4.columns.get_loc("ед. изм.") + 1)
    volume_col = get_column_letter(df_form4.columns.get_loc("Объем единицы, м3") + 1)
    final_volume_col = get_column_letter(df_form4.columns.get_loc("Итоговый объем, м3") + 1)
    date_col = get_column_letter(df_form4.columns.get_loc("Приведенная дата") + 1)

    # Заполняем формулы
    for row in range(2, len(df_form4) + 2):
        # Формула для столбца "Объем единицы, м3" с внешней ссылкой
        volume_formula = f'=IFERROR(VLOOKUP({code_col}{row}, \'[{os.path.basename(form1_filepath)}]Sheet1\'!A:L, 12, 0), \'[{os.path.basename(form1_filepath)}]Sheet1\'!$Q$6)'
        sheet[f"{volume_col}{row}"].value = volume_formula

        # Формула для столбца "Количество с учетом единицы измерения"
        quantity_formula = f'=IF({unit_col}{row}="шт", {start_quantity_col}{row}, CEILING({start_quantity_col}{row}, 1))'
        sheet[f"{quantity_col}{row}"].value = quantity_formula

        # Формула для столбца "Итоговый объем, м3"
        final_volume_formula = f"={volume_col}{row}*{quantity_col}{row}"
        sheet[f"{final_volume_col}{row}"].value = final_volume_formula

        # Обновление прогресса
        progress_var.set(f"Обработка данных: {row - 1} / {len(df_form4)} строк")
        root.update()

    # Создаем лист "СВОД"
    sheet_svod = workbook.create_sheet("СВОД")

    # Добавляем заголовки
    sheet_svod["A1"] = "Приведенная дата"
    sheet_svod["B1"] = "Объем, м3"
    sheet_svod["C1"] = "Количество, строк"
    sheet_svod["D1"] = "Количество, штук"
    sheet_svod["E1"] = "Количество документов, штук"

    # Получаем уникальные приведенные даты
    unique_dates = df_form4["Приведенная дата"].dropna().unique()

    # Подсчет уникальных кодов товаров и заполнение данных
    for idx, date in enumerate(sorted(unique_dates), start=2):
        sheet_svod[f"A{idx}"].value = date

        # Формула для суммы объемов по приведенной дате
        date_cell = f"A{idx}"
        sheet_svod[
            f"B{idx}"] = f"=SUMIF('{sheet.title}'!{date_col}:{date_col}, {date_cell}, '{sheet.title}'!{final_volume_col}:{final_volume_col})"

        # Подсчет количества строк для текущей даты
        row_count = df_form4[df_form4["Приведенная дата"] == date].shape[0]
        sheet_svod[f"C{idx}"].value = row_count

        # Формула для суммы количества с учетом единицы измерения
        sheet_svod[
            f"D{idx}"] = f"=SUMIF('{sheet.title}'!{date_col}:{date_col}, {date_cell}, '{sheet.title}'!{quantity_col}:{quantity_col})"

        # Подсчет количества уникальных документов для текущей даты
        document_count = df_form4[df_form4["Приведенная дата"] == date]["Номер документа"].nunique()
        sheet_svod[f"E{idx}"].value = document_count

    # Сохранение и закрытие файла
    try:
        workbook.save(new_filepath)
        workbook.close()
    except Exception as e:
        messagebox.showerror("Ошибка сохранения", f"Не удалось сохранить файл. Ошибка: {e}")
        return

    # Сообщаем об успешной обработке
    messagebox.showinfo("Успешно", f"Файл формы 4 успешно обработан и сохранен по пути: {new_filepath}")
    #on_form4_done()
